import openpyxl
import config
import requests
from bs4 import BeautifulSoup
from lxml import  html
total_page = config.last_page_number
def write_list_to_excel(filename, data):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        headers_row = ['name','ISBN', 'authors', 'publish', 'imgSrc','price', 'numberOfPages', 'Dimensions', 'description']
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers_row)
    for new_row in data:
        ws.append(new_row)
    wb.save(filename)
def get_column(file_name):
    workbook = openpyxl.load_workbook(filename=file_name)
    worksheet = workbook.active
    ans = []
    for i in range(10,10776):
        cell_name = f'D{i}'
        cell_value = worksheet[cell_name].value
        print(cell_value)
    return ans

page = 1
while page <= 18:
    url = "https://www.qbd.com.au/the-169-storey-treehouse/andy-griffiths-terry-denton/9781760987855/"
    print(url)
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'lxml')
    print(soup)
    break
    tree = html.fromstring(r.content)
    detail_links = tree.xpath('//div[@class="q2k21ptc"]/a')

    print(detail_links)
    all_data = []
    for link in detail_links:

        new_url = 'https://www.qbd.com.au' + link.attrib['href']
        print(new_url)
        res = requests.get(new_url)

        tree = html.fromstring(res.content)
        soup = BeautifulSoup(res.text, 'lxml')
        try:
            name = tree.xpath('//h1[@itemprop="name"]')[0].text.strip()
        except:
            name = ""
        try:
            authors=tree.xpath('//div[@class="line clear"]/span')[0].text_content().strip()
        except:
            authors = ""
        try:
            description = soup.find('div', class_='readmore').text.strip()
        except:
            description = ""
        try:
            ISBN = tree.xpath('//span[@itemprop="isbn"]')[0].text.strip()
        except:
            ISBN = ""
        try:
            price = soup.find('span', class_='strikethrough').text.strip()
        except:
            price = ""
        try:
            publish = tree.xpath('//span[@itemprop="datePublished"]')[0].text.strip()
        except:
            publish = ""
        try:
            Binding = tree.xpath('//span[@itemprop="bookFormat"]')[0].text.strip()
        except:
            Binding = ""
        try:
            numberOfPages = tree.xpath('//span[@itemprop="numberOfPages"]')[0].text.strip()
        except:
            numberOfPages = ""
        try:
            Dimensions = tree.xpath('//div[@class="line"]')[-1].text_content().replace('Dimensions', "").strip()
        except:
            Dimensions = ""
        try:
            imgSrc = tree.xpath('//img[@itemprop="image"]')[0].attrib['src']
        except:
            imgSrc = ""



        data_arr = [name, ISBN, authors, publish, imgSrc,price, numberOfPages, Dimensions, description]
      
        print(data_arr)
        all_data.append(data_arr)
    write_list_to_excel('./output.xlsx', all_data)
    page += 1